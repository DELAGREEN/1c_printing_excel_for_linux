###!!!!!!!!minimum viable product

import sys
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook
import time
import os
from errors_logger import ErrorsLogger as Logger
#from icecream import ic


class textcolors:
    HEADER = '\033[95m'
    BLUE = '\033[94m'
    CYAN = '\033[96m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    END = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

param_list = f'''
{textcolors.YELLOW}-excel {textcolors.GREEN}активирует запись в EXCEL, принимает в качестве параметров {textcolors.BLUE}[путь к excel файлу][путь к txt файлу][разделитель1][разделитель2] 

'''

doc = f'''
{textcolors.YELLOW}-excel {textcolors.BLUE}[путь к excel файлу][путь к txt файлу][разделитель1][разделитель2]
        
{textcolors.YELLOW}-excel {textcolors.GREEN}флаг который показывает какого типа файл нужно будет собрать {textcolors.END}
{textcolors.YELLOW}[путь к excel файлу] {textcolors.GREEN} Путь куда конкретно сохранить файл Excel с расширением {textcolors.BLUE}(.xlsx)
{textcolors.YELLOW}[путь к txt файлу] {textcolors.GREEN} Путь к txt файлу сформированный 1с в котором лежат матаданные сериализатора
{textcolors.YELLOW}[разделитель1] {textcolors.GREEN} Разделитель между данными по одной конкретной ячейки {textcolors.BLUE}[номер строки][разделитель1][номер столбца][разделитель1][значение ячейки]
{textcolors.YELLOW}[разделитель2] {textcolors.GREEN} Разделитель между данными ячеек {textcolors.BLUE}[номер строки][разделитель1][номер столбца][разделитель1][значение ячейки][разделитель2][номер строки][разделитель1][номер столбца][разделитель1][значение ячейки]

'''

def is_empty_excel_or_txt(_path_to_excel:str, _path_to_txt:str):

    '''Проверяем создан файл если нет создаем'''
    if not os.path.exists(_path_to_excel):
        wb = Workbook()                      
        wb.save(_path_to_excel)

    '''Проверяем создан файл если нет создаем'''
    if not os.path.exists(_path_to_txt):
        Logger.print_error('Файл с расширением .txt не обнаружен')
        print(f'{textcolors.YELLOW}Ошибка. Файл с расширением .txt не обнаружен')


def read_txt(_path_to_txt:str):
    
    with open(_path_to_txt, 'r') as txt_file:
        string = txt_file.read()
        #Вычищаем возможную помойку
        try:
            string = string.replace('\ufeff', '')   #1С бывает засовывает свою какую то параметризацию выглядик как вот эта какаха которую я отрезаю
        finally:
            return string
  

'''
Функция принимает строку row:column:value,row:column:value,row:column:value 
дробит ее на параметры принимаемые openpyxl 
НАПРИМЕР:row:column:value для поячеечной записи в excel файл
'''
def writer_for_excel(_strings, path_to_excel:str, separator_first:str, separator_end:str):

    #дробим строку вида [номер строки, номер столбца, значение ячейки][разделитель][номер строки, номер столбца, значение ячейки]
    list_strings = _strings.split(separator_end)                             #и получаем [номер строки][разделитель][номер столбца][разделитель][значение ячейки]                     
    
    '''Удаляем задолбавший \n в конце строки'''
    '''Требует ОСОБОГО ВНИМАНИЯ'''
    '''Возможны Баги в перспективе'''
    if list_strings[-1][-1] == '\n':
        list_strings[-1] = list_strings[-1][:-1]
    
    '''Требует Особого Внимания'''
    wb = load_workbook(path_to_excel)                                        #Передаем классу, методу класса Файл который нужно открыть
    
    for i in list_strings:
        
        cell_data_set = i.split(separator_first)                             #дробим строку на [номер строки][номер столбца][значение ячейки]

        match cell_data_set:
            case _row, _column, _value:                                      #дробим строку на [номер строки][номер столбца][значение ячейки] 
                #_row, _column, _value = cell_data_set                       #Эта строка аналогична верхней, более читабельна, оставил что бы не забыть как это выглядит по нормальному
                ws = wb.worksheets[0]                                        #Номер страницы для записи
                #ws[column] = value                                          #colunm должен равняться номеру столбца НАПРИМЕР: А1
                ws.cell(row=int(_row), column=int(_column)).value = _value   #передаем текст в ячейку 
        
            case _row, _column, _value, _sheet_num:                          #дробим строку на [номер строки][номер столбца][значение ячейки][номер страницы]
                #_row, _column, _value, _sheet_num = cell_data_set             
                ws = wb.worksheets[int(_sheet_num)]                          #Номер страницы для записи
                ws.cell(row=int(_row), column=int(_column)).value = _value   #передаем текст в ячейку

            case _:
                Logger.print_error('Не правильно сереализирован TXT файл.')
                print('Не правильно сереализирован TXT файл.')

    wb.save(path_to_excel)                                                   #сохраняем файл
    Logger.print_info('Write is сomplite!')
    print(f'{textcolors.YELLOW}Write is сomplite!')

def parse_param():
    
    param_name = 'None'
    path_to_excel = 'None'
    path_to_txt = 'None'
    separator_first = 'None'
    separator_end = 'None'

    param_name = sys.argv[1]
    match param_name:
        case '--excel'| '-excel':                                            #Если есть тригер выполняем код ниже           
            try:
                path_to_excel = str(sys.argv[2])                             #Захватываем путь к excel
                path_to_txt = str(sys.argv[3])                               #Захватываем путь к txt
                separator_first = str(sys.argv[4])                           #Захватываем разделитель между значениями ячек и данными int[4]int[4]value
                separator_end = str(sys.argv[5])                             #Захватываем разделитель между данными одной ячейки int[4]int[4]value[5]int[4]int[4]value[5] и тд
                is_empty_excel_or_txt(path_to_excel, path_to_txt)
                writer_for_excel(read_txt(path_to_txt), path_to_excel, separator_first, separator_end)

            except Exception as _ex:
                if str(_ex) == 'list index out of range':
                    Logger.print_error(f'Индекс списка вне диапазона. Колличество страниц в excel не совпадает с колличеством которое пытаюся записать!\n'
                                    f'param_name = {param_name}\n'
                                    f'path_to_excel = {path_to_excel}\n'
                                    f'path_to_txt = {path_to_txt}\n'
                                    f'separator_first = {separator_first}\n'
                                    f'separator_end = {separator_end}\n'
                                    f'EXCEPTION: {_ex}')
                    print(f'{textcolors.RED}Ошибка.{textcolors.YELLOW} Индекс списка вне диапазона. Обратитесь к Администратору!')
            
                else:
                    Logger.print_error(f'Неизвестная ошибка: Проверьте правильность параметра:'
                                    f'param_name = {param_name}\n'
                                    f'path_to_excel = {path_to_excel}\n'
                                    f'path_to_txt = {path_to_txt}\n'
                                    f'separator_first = {separator_first}\n'
                                    f'separator_end = {separator_end}\n'
                                    f'EXCEPTION: {_ex}') 
                    print(f'{textcolors.RED}Ошибка.{textcolors.YELLOW} Неизвестная ошибка: Проверьте правильность параметра:{param_name}\n{_ex}')
        
        case '--help'|'-help':
            print(f'{textcolors.CYAN}***HELP SHEET***\n{doc}')  
   
        case _:
            Logger.print_error(f'Неизвестный параметр: {param_name}')
            print(f'{textcolors.RED}Ошибка. {textcolors.YELLOW}Неизвестный параметр: {param_name}')
            

if __name__ == '__main__':

    Logger.print_info('Start programm...')

    match len(sys.argv):

        case 1:
            print(f'{textcolors.RED}Внимание! {textcolors.YELLOW}Дополнительные параметры не заданы.\n' 
                'Модуль не работает без дополнительных параметров.\n'
                f'В качестве параметров принимается {textcolors.GREEN}Строка(String){textcolors.YELLOW} в качестве разделителей параметров {textcolors.GREEN}Пробел(Space).{textcolors.YELLOW} Пример: {textcolors.BLUE}[Параметр1] [Параметр2] [Параметр3]{textcolors.YELLOW} и тд (без квадратных скобок).\n'
                '\n'
                'На данный момент принимаются:'
                f'{param_list}\n'
                f'{textcolors.YELLOW}*ИЛИ*\n'
                'Ознакомтесь с документацией c помошью команды -help.'
                f'{textcolors.END}')
            Logger.print_warning('Дополнительные параметры не заданы.')
    
        case 2:
            parse_param()   #Парсим параметры

        case 6:
            parse_param()   #Парсим параметры 
        
        case _:
            print(f'{textcolors.RED}Ошибка. {textcolors.YELLOW}Слишком мало параметров. Воспользуйтесь командой -help.')
            Logger.print_error('Длина массива параметров не соответствует требуемым.')

Logger.print_info('End programm...')
sys.exit(1)